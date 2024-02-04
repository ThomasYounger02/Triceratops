
# 导入
from openpyxl import load_workbook, Workbook

# 通过文件路径，打开已有工作簿
wb1 = load_workbook('./demo_excel.xlsx')
# 用 Workbook() 创建新工作簿
wb2 = Workbook()

from openpyxl import Workbook

# 新建工作簿
new_wb = Workbook()
# 将新建的工作簿保存为【new_excel.xlsx】
new_wb.save('./new_excel.xlsx')

# 活动工作表
from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

# 按照表名取表
from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 按表名取表
fhy_ws = staff_wb['上半年公司名单']  # fhy为first half year（上半年）的缩写
shy_ws = staff_wb['下半年公司名单']  # shy为second half year（下半年）的缩写

# 打印工作簿对象
print(staff_wb)
# 打印工作表对象
print(fhy_ws)
print(shy_ws)



# 打印获取到的第五行数据
print(active_ws[5])
# 打印获取到的第二列数据
print(active_ws['B'])


# 输出结果确实为元组，但元组中的每一个元素均为<Cell '工作表名称'.坐标>的形式


from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

# 返回第2行至第12行，第2列（B列）至第3列（C列）这个范围的单元格内的所有数据（值）
for row in fhy_ws.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3, values_only=True):
    print(row)

from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./公司人员名单.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

info_list = ['S1911', '萧爵瑟', 3000, '内容']
info_tuple = ('S1912', '吴琐薇', 5000, '销售')

active_ws.append(info_list)
active_ws.append(info_tuple)

# 保存工作簿为【append_demo.xlsx】
staff_wb.save('./append_demo.xlsx')


from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./公司人员名单.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

info_list = ['S1911', '萧爵瑟', 3000, '内容']
info_tuple = ('S1912', '吴琐薇', 5000, '销售')

active_ws.append(info_list)
active_ws.append(info_tuple)

# 保存工作簿为【append_demo.xlsx】
staff_wb.save('./append_demo.xlsx')

# method 1
from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']


from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

# 返回第2行至第12行，第2列（B列）至第3列（C列）这个范围的单元格内的所有数据（值）
for row in fhy_ws.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3, values_only=True):
    print(row)


from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

# 返回第2行至第12行，第2列（B列）至第3列（C列）这个范围的单元格内的所有数据（值）
for row in fhy_ws.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3, values_only=True):
    print(row)


from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./公司人员名单.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

info_list = ['S1911', '萧爵瑟', 3000, '内容']
info_tuple = ('S1912', '吴琐薇', 5000, '销售')

active_ws.append(info_list)
active_ws.append(info_tuple)

# 保存工作簿为【append_demo.xlsx】
staff_wb.save('./append_demo.xlsx')





from openpyxl import load_workbook, Workbook

# 打开【10月考勤统计.xlsx】工作簿
wb = load_workbook('./material/10月考勤统计.xlsx')
# 获取活动工作表
ws = wb.active

# 获取表头
late_header = []
for cell in ws[1]:
    late_header.append(cell.value)

# 新建工作簿
new_wb = Workbook()
# 获取新工作簿中的工作表
new_ws = new_wb.active

# 将表头写入新工作簿的工作表中
new_ws.append(late_header)

# 从第二行开始遍历表格
for row in ws.iter_rows(min_row=2, values_only=True):
    # 取出姓名，迟到时间和迟到次数
    name = row[1]
    time = row[3]
    number = row[-1]
    # 判断是否迟到
    if time > 45 and number > 3:
        print('{}迟到了{}分钟，迟到了{}次'.format(name, time, number))
        # 将迟到人员信息写入新工作簿的工作表中
        new_ws.append(row)

# 将新工作簿保存为【10月迟到人员信息.xlsx】
new_wb.save('./material/10月迟到人员信息.xlsx')



from openpyxl import load_workbook

# 打开工作簿【10月考勤统计.xlsx】，获取活动工作表
wb = load_workbook('./material/10月考勤统计.xlsx')
ws = wb.active

# 创建迟到人员字典
info_dict = {}

# 循环读取除表头外的表格数据
for row in ws.iter_rows(min_row=2, values_only=True):
    # 取出员工工号
    staff_id = row[0]
    # 取出迟到次数
    staff_late = row[-1]
    # 将信息添加入字典，字典格式为{'员工工号': '迟到次数'}
    info_dict[staff_id] = staff_late

# 打开工作簿【迟到次数月度统计（10月更新）.xlsx】，获取活动工作表
monthly_wb = load_workbook('./material/迟到次数月度统计（10月更新）.xlsx')
monthly_ws = monthly_wb.active

# 循环读取出表头外的表格数据
for monthly_row in monthly_ws.iter_rows(min_row=3, max_col=13, values_only=True):
    # 取出员工工号
    member_id = monthly_row[0]
    # 取出十月份的迟到次数
    member_late = monthly_row[-1]
    # 匹配迟到次数是否相等
    if member_late != info_dict[member_id]:
        print('工号{}迟到情况不匹配，请核查后更新'.format(member_id))


# 工作簿
from openpyxl import load_workbook, Workbook

# 通过文件路径，打开已有工作簿
wb1 = load_workbook('./demo_excel.xlsx')
# 用 Workbook() 创建新工作簿
wb2 = Workbook()


from openpyxl import Workbook

# 新建工作簿
new_wb = Workbook()
# 将新建的工作簿保存为【new_excel.xlsx】
new_wb.save('./new_excel.xlsx')

from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 按表名取表
fhy_ws = staff_wb['上半年公司名单']  # fhy为first half year（上半年）的缩写
shy_ws = staff_wb['下半年公司名单']  # shy为second half year（下半年）的缩写

# 打印工作簿对象
print(staff_wb)
# 打印工作表对象
print(fhy_ws)
print(shy_ws)

# 工作表
# method 1
from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

# 返回第2行至第12行，第2列（B列）至第3列（C列）这个范围的单元格内的所有数据（值）
for row in fhy_ws.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3, values_only=True):
    print(row)

from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./公司人员名单.xlsx')
# 获取活动工作表
active_ws = staff_wb.active

info_list = ['S1911', '萧爵瑟', 3000, '内容']
info_tuple = ('S1912', '吴琐薇', 5000, '销售')

active_ws.append(info_list)
active_ws.append(info_tuple)

# 保存工作簿为【append_demo.xlsx】
staff_wb.save('./append_demo.xlsx')

# 单元格
# method 1
from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取【'上半年公司名单'】工作表
fhy_ws = staff_wb['上半年公司名单']

from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取活动工作表
staff_ws = staff_wb.active

# for循环遍历，取出第三行的所有单元格对象
for row_cell in staff_ws[3]:
    print(row_cell)

# for循环遍历，取出第三列（C列）的所有单元格对象
for col_cell in staff_ws['C']:
    print(col_cell)


from openpyxl import load_workbook

# 打开【公司人员名单.xlsx】工作簿
staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
# 获取活动工作表
staff_ws = staff_wb.active

# 打印单元格对象A1
print(staff_ws['A1'])


# 修改单元格对象C2的值为10000
staff_ws['C2'].value = 10000

# 打印修改后的单元格对象C2的值
print(staff_ws['C2'].value)


########################################################3
# 表格读写
